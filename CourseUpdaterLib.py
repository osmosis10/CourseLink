'''
Author: Moses Lemma

Description: courseUpdaterLib.py is a small module containing the functions used to perform various steps of preparing
             the course data to be uploaded to SAMI.

             Uses the openpyxl library for all excel methods.
'''
# IMPORTS
import os
import subprocess
import sys
import time

# Try to install openpyxl, if not installed, pip will install
try: 
    import openpyxl

except:
    libs = ["openpyxl"]
    for lib in libs:
        subprocess.check_call([sys.executable, "-m", "pip", "install", lib])
        print(f"Installed {lib} libary\n")
        time.sleep(3)



def checkIfAllFilesPresent():
    filesList = ["report2", 
             "report1", "report3"]
    count = 0

    curdir = os.getcwd()
    dirList = os.listdir(curdir)

    for fileName in dirList:
        if (fileName.endswith(".xlsx")):
            for prefix in filesList:
                if (fileName.startswith(prefix)):
                    count += 1
            if (count == 3):
                return True
    return False

# Additional check to ensure file exists in folder
def getFileName(prefix):
    curdir = os.getcwd()
    fileName = ""
    for files in os.listdir(curdir):
        if (files.startswith(prefix)):
            fileName = files
    
    if (not fileName):
        print(f"{prefix} not found !")
        return False
    
    return fileName

def getSheet(fileName):
    report =  openpyxl.load_workbook(fileName)

    Sheets = report.sheetnames
    NameofSheet = Sheets[0]
    return report[NameofSheet], report

# (Used in Option 1). Obtain all the providers in a dictionary with key-value
#                     pairs from Report #3
def getAllProviders():
    report3Name = "report3.xlsx"
    fileName = getFileName(report3Name)
    if (fileName == False):
        return False # File not found

    try:
        Sheet, report = getSheet(fileName)
        providerDict = {} # Get key-value pairs for 'course title' and 'training provider'

        for row in (Sheet.iter_rows(min_row=14, min_col=1, max_col=4,values_only=True)):
            providerDict[list(row)[0]] = list(row)[3]

        return providerDict
    
    except Exception as e:
        with open('debug_log.txt', 'w') as f:
            f.write(str(e))




# Option 1. Updates the providers in report 2 with the 'Developed by' values
#           from report 3 (if they exist).
def updateReport2(providerDict):
    report2Name = "report2.xlsx " # Report 2
    fileName = getFileName(report2Name)
    if (fileName == False):
        return False # File not found

    try:
        Sheet, report2 = getSheet(fileName) # selecting the first sheet if more than 1 is available
        updateFlag, outputList, count = 0, [], 0

        # Iterate through each row at a time with 8 columns at a time
        for row in (Sheet.iter_rows(min_row=19, min_col=1, max_col=8)):
            # Get the cell values of the first 8 columns in one list 
            rowList = [cell.value for cell in row]

            if (rowList[4] in providerDict and rowList[7] == "Value"):
                providerName = providerDict[rowList[4]] # store provider name

                outputList.append(f"Updating {rowList[1]} {rowList[2]}'s provider, {rowList[7]}, with {providerName}")
                row[7].value = providerName # Update the cell value
                updateFlag = 1 # set flag to 1
                count += 1
                
        outputList.append(f"\nTOTAL: {count}")
        report2.save(fileName)
        return updateFlag, outputList
    
    except Exception as e:
        print(e)
        with open('debug_log.txt', 'w') as f:
            f.write(str(e))


# Option 2. Updating the training status for the LinkedIn Learning in report #1
def updateReport1TrainingStatus():
    report1Name = "report1.xlsx" # Report 1

    fileName = getFileName(report1Name)
    if (fileName == False):
        return False # File not found
    
    Sheet, report1 = getSheet(fileName) # selecting the first sheet if more than 1 is available
    updateFlag, outputList, count = 0, ["Updating LinkedIn Learning Status For:\n\n"], 0

    try:

        # Iterate through each row at a time and 8 columns at a time
        for row in (Sheet.iter_rows(min_row=19, min_col=1, max_col=8)):
            rowList = [cell.value for cell in row]

            if (str(row[7].value).lower().replace(" ", "").startswith("linkedin", 0, 8) and row[5].value == "Approved"):
                print(f"Value = {row[7].value.lower().replace(" ", "")}")

                count += 1
                row[5].value = "Completed" # Perform update
                output = f"{count}. {rowList[1]} {rowList[2]}\n\n"
                outputList.append(output)
                updateFlag = 1
                

        outputList.append(f"\nTOTAL: {count}")
        report1.save(fileName)
        return updateFlag, outputList
    
    except Exception as e:
        with open('debug_log.txt', 'w') as f:
            f.write(str(e))


# Option 3. Checking for duplicate records in report #2
def checkForDupesinReport2():
    report2Name = "report2.xlsx" # Report 2
    fileName = getFileName(report2Name)
    if (fileName == False):
        return False # File not found

    Sheet, report2 = getSheet(fileName) # selecting the first sheet if more than 1 is available

    rowDict = {} # dictionary structured as {key = unique record details: int number of occurences}
    updateFlag, outputList, count = 0, [], 0
    
    for row1 in (Sheet.iter_rows(min_row=19, min_col=1, max_col=10)):
        rowList1 = [cell.value for cell in row1]
        
        # Key used to ensure uniqueness
        key = (f"{rowList1[0]}, {rowList1[1]}, {rowList1[2]}, {rowList1[3]}, {rowList1[4]}," 
               f"{rowList1[5]}, {rowList1[6]}, {rowList1[7]}, {rowList1[8]}, {rowList1[9]}")
        
        if (key not in rowDict):
            rowDict[key] = 0

        elif (key in rowDict):
            rowDict[key] = rowDict[key] + 1
            count += 1
        
        if (rowDict[key] == 1):
            dupe = f"{rowList1[1]} {rowList1[2]}"
            outputList.append(dupe)
            updateFlag = 1

    outputList.append(f"\nTOTAL: {count}")
    return updateFlag, outputList # 0 if no dupes found, 1 if at least one found


# Option 4. (Previously removeArchive) Displays all course titles that start with "Archive"
#           and contain '::'
def displayArchive():
    report2Name = "report2" # Report 2

    fileName = getFileName(report2Name)
    if (fileName == False):
        return False # File not found
    
    Sheet, report2 = getSheet(fileName) # selecting the first sheet if more than 1 is available
    updateFlag, outputList, count = 0, [], 0

    try: 
        # Iterate through each row at a time and 8 columns at a time
        for row in (Sheet.iter_rows(min_row=19, min_col=1, max_col=8)):
            rowList = [cell.value for cell in row]

            if (rowList[4].startswith("Archive ") and "::" in rowList[4]):

                output = f"{rowList[1]} {rowList[2]}\n"
                outputList.append(output)
                updateFlag = 1
                count += 1
        
        outputList.append(f"\nTOTAL: {count}")

        report2.save(fileName)
        return updateFlag, outputList

    except Exception as e:
        with open('debug_log.txt', 'w') as f:
            f.write(str(e))